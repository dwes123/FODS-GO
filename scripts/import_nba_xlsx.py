#!/usr/bin/env python3
"""
Reads "The Dynasty Association.xlsx" and emits SQL to populate fantasy_basketball_db.

USAGE
    python scripts/import_nba_xlsx.py "C:/path/to/The Dynasty Association.xlsx" > /tmp/nba_import.sql
    docker exec -i fantasy_postgres psql -U admin -d fantasy_basketball_db < /tmp/nba_import.sql

The script outputs one big idempotent SQL transaction. It TRUNCATEs the players table
first (so re-running gives a clean state) but preserves teams, agencies, league_settings,
and key_dates.

What it imports:
  - 30 team financial snapshots (Franchise Financial Status tab) → teams.cap_space, etc.
  - All player rows from each team tab → players (name, position, real_life_team, contracts,
    contract annotations from cell colors, on_g_league flag for G-League section)
  - Fantrax IDs from "Player Teams" tab → players.fantrax_id (matched by name + real_life_team)
  - Agency assignments from "Agent Directory" tab → players.agency_id
  - Cap thresholds from "League Office" tab → league_settings (MLE, TPMLE, BAE, etc.)
  - Renames the seeded agencies to the actual agency names

Cell colors map to contract tags via this legend (from League Office, col B):
    FFFF0000 = Team Option           FFFFFF00 = Qualifying Offer Extended
    FFFF9900 = Player Option         FF00FF00 = UFA Year
    FFFF00FF = Qualifying Offer      FF8E7CC3 = 12/1 Trade Restriction
    FF0000FF = DPE Designation
"""

import sys
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl required: pip install openpyxl\n")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

LEAGUE_ID = "55555555-5555-5555-5555-555555555555"

# Team-tab name → our canonical 3-letter abbreviation in fantasy_basketball_db.teams.
TEAM_TAB_TO_ABBREV = {
    "ATL": "ATL", "BOS": "BOS", "BRK": "BKN", "CHA": "CHA", "CHI": "CHI",
    "CLE": "CLE", "DAL": "DAL", "DEN": "DEN", "DET": "DET", "GS":  "GSW",
    "HOU": "HOU", "IND": "IND", "LAC": "LAC", "LAL": "LAL", "MEM": "MEM",
    "MIA": "MIA", "MIL": "MIL", "MIN": "MIN", "NOP": "NOP", "NYK": "NYK",
    "OKC": "OKC", "ORL": "ORL", "PHI": "PHI", "PHX": "PHX", "POR": "POR",
    "SAC": "SAC", "SAS": "SAS", "TOR": "TOR", "UTA": "UTA", "WSH": "WAS",
}

# "Player Teams" tab uses different shorthand (NO/NY/PHO/SA/WAS instead of NOP/NYK/PHX/SAS).
# Used for matching players from that tab back to our team UUIDs.
PLAYER_TEAMS_TO_ABBREV = {
    "ATL": "ATL", "BKN": "BKN", "BOS": "BOS", "CHA": "CHA", "CHI": "CHI",
    "CLE": "CLE", "DAL": "DAL", "DEN": "DEN", "DET": "DET", "GS":  "GSW",
    "HOU": "HOU", "IND": "IND", "LAC": "LAC", "LAL": "LAL", "MEM": "MEM",
    "MIA": "MIA", "MIL": "MIL", "MIN": "MIN", "NO":  "NOP", "NY":  "NYK",
    "OKC": "OKC", "ORL": "ORL", "PHI": "PHI", "PHO": "PHX", "POR": "POR",
    "SA":  "SAS", "SAC": "SAC", "TOR": "TOR", "UTA": "UTA", "WAS": "WAS",
    "(N/A)": None,  # historical / unowned
}

# Real-life team field on player rows in team tabs uses a third variant. Same mapping logic.
# Build a superset that handles every variant we've seen.
ALL_TEAM_ALIASES = {**TEAM_TAB_TO_ABBREV, **PLAYER_TEAMS_TO_ABBREV}

# Cell fill color (ARGB hex) → contract tag.
COLOR_TO_TAG = {
    "FFFF0000": "Team Option",
    "FFFF9900": "Player Option",
    "FFFF00FF": "Qualifying Offer",
    "FFFFFF00": "Qualifying Offer Extended",
    "FF00FF00": "UFA Year",
    "FF8E7CC3": "12/1 Trade Restriction",
    "FF0000FF": "DPE Designation",
}

# Tags that ANNOTATE a cell (i.e., decorate alongside a dollar amount). Stored in
# players.contract_annotations JSONB. Everything else is the primary status of the year
# and goes into the contract_YYYY column directly.
ANNOTATION_TAGS = {"12/1 Trade Restriction", "DPE Designation"}

# Agency names from the spreadsheet → seeded agency UUIDs from migration 003.
AGENCY_NAME_TO_ID = {
    "Billy the Kid Sports Agency": "a9e0c111-0000-0000-0000-000000000001",
    "The Boys Sports Agency":      "a9e0c222-0000-0000-0000-000000000002",
}

# Contract column years on team tabs go from 2026-27 (col F=6) through 2036-37 (col P=16).
# Maps to our contract_2026..contract_2036 columns.
CONTRACT_START_COL = 6
CONTRACT_FIRST_YEAR = 2026
CONTRACT_NUM_YEARS = 11


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sql_str(v):
    """Quote a string value for SQL, handling None and embedded single quotes."""
    if v is None or v == "":
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def sql_num(v):
    if v is None or v == "":
        return "NULL"
    return str(v)


def normalize_name(name):
    """Collapse whitespace and strip — used for matching across tabs."""
    if not name:
        return ""
    return re.sub(r"\s+", " ", str(name)).strip()


def cell_color(cell):
    """Return the ARGB fg color string of a cell, or None if no fill."""
    fill = cell.fill
    if not fill or fill.patternType is None:
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    rgb = fg.rgb
    # openpyxl returns either an ARGB string like 'FFFF0000' or a theme/index ref.
    if isinstance(rgb, str) and len(rgb) == 8 and rgb != "00000000":
        return rgb.upper()
    return None


def normalize_position(pos):
    """Uppercase and trim a 'G,F' / 'G' / 'F,C' style position string."""
    if not pos:
        return None
    parts = [p.strip().upper() for p in str(pos).split(",") if p.strip()]
    return ",".join(parts) if parts else None


# ---------------------------------------------------------------------------
# Extractors
# ---------------------------------------------------------------------------

def extract_team_roster(ws, team_tab, team_abbrev):
    """Yield dicts for each player row in a team tab."""
    in_nba = False
    in_gleague = False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        # row is a tuple of Cell objects (NOT values_only here — we need cell.fill)
        first = row[0].value if row else None
        if first == "NBA ROSTER SALARY INFORMATION":
            in_nba, in_gleague = True, False
            continue
        if first == "G-LEAGUE ROSTER SALARY INFORMATION":
            in_nba, in_gleague = False, True
            continue
        if not (in_nba or in_gleague):
            continue
        if len(row) < 5:
            continue

        transaction = row[0].value
        notes = row[1].value if len(row) > 1 else None
        position = row[2].value if len(row) > 2 else None
        name = row[3].value if len(row) > 3 else None
        real_life_team = row[4].value if len(row) > 4 else None

        if not name or not str(name).strip():
            continue
        if str(name).strip().lower() == "name":  # header row
            continue
        # Skip junk placeholder rows that have a name but no position AND no real_life_team
        # (these appear in some G-League sections as draft-slot placeholders).
        if not position and not real_life_team:
            continue

        # Extract contract values + colors
        contracts = {}     # year (int) → string ("$XXX" or "UFA Year" or "G-League Contract")
        annotations = {}   # year (str) → list[str] of annotation tags
        for offset in range(CONTRACT_NUM_YEARS):
            col_idx = CONTRACT_START_COL - 1 + offset  # 0-based
            if col_idx >= len(row):
                break
            cell = row[col_idx]
            v = cell.value
            color = cell_color(cell)
            tag = COLOR_TO_TAG.get(color)
            year = CONTRACT_FIRST_YEAR + offset

            if v is None and tag is None:
                continue

            if isinstance(v, (int, float)):
                amt = int(round(v))
                if amt <= 0:
                    # $0 cells are typically placeholders (empty future years).
                    # Preserve any tag (option / QO / etc.) so the annotation isn't lost,
                    # but don't write a "$0" contract value.
                    if tag:
                        annotations.setdefault(str(year), []).append(tag)
                    continue
                contracts[year] = f"${amt}"
                if tag:
                    annotations.setdefault(str(year), []).append(tag)
            elif isinstance(v, str):
                txt = v.strip()
                if txt.upper() == "UFA":
                    contracts[year] = "UFA Year"
                elif txt.lower() == "g-league contract":
                    contracts[year] = "G-League Contract"
                elif txt:
                    # Some other text — store as-is
                    contracts[year] = txt
                if tag and tag in ANNOTATION_TAGS:
                    annotations.setdefault(str(year), []).append(tag)
            elif tag:
                # No value but a color (rare)
                if tag in ANNOTATION_TAGS:
                    annotations.setdefault(str(year), []).append(tag)
                else:
                    contracts[year] = tag

        yield {
            "team_abbrev": team_abbrev,
            "transaction": str(transaction).strip() if transaction else None,
            "notes": str(notes).strip() if notes else None,
            "position": normalize_position(position),
            "name": normalize_name(name),
            "real_life_team": str(real_life_team).strip() if real_life_team else None,
            "contracts": contracts,
            "annotations": annotations,
            "on_g_league": in_gleague,
            # Two-way contracts: per CBA the "Latest Financial Transaction" is "Two-Way".
            "on_two_way": str(transaction).strip().lower() == "two-way" if transaction else False,
        }


def extract_player_teams(ws):
    """Yield (fantrax_id, name, real_life_team_abbrev_canonical, position) from Player Teams tab."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or row[0] is None:
            continue
        fid, name, team, pos = row[:4]
        if not fid or not name:
            continue
        team_canonical = PLAYER_TEAMS_TO_ABBREV.get(str(team).strip()) if team else None
        yield {
            "fantrax_id": str(fid).strip(),
            "name": normalize_name(name),
            "real_life_team": team_canonical,
            "position": normalize_position(pos),
        }


def extract_agency_assignments(ws):
    """Yield (player_name, agency_name) from Agent Directory tab.
    Skips header and banner rows; uses col H (idx 7) for the agency name."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            continue
        if not row or len(row) < 8:
            continue
        name = row[0]
        agency = row[7]
        if not name or not agency:
            continue
        if "agency" not in str(agency).lower():
            continue
        # Skip header rows ("PLAYER" etc.)
        if str(name).strip().upper() in ("PLAYER", ""):
            continue
        yield {"name": normalize_name(name), "agency": str(agency).strip()}


def extract_team_financials(ws):
    """Yield team financial snapshots from Franchise Financial Status tab."""
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or len(row) < 11:
            continue
        # Col B (idx 1) = team abbrev
        abbrev = row[1]
        if not abbrev or str(abbrev).strip() not in TEAM_TAB_TO_ABBREV:
            continue
        canonical = TEAM_TAB_TO_ABBREV[str(abbrev).strip()]
        yield {
            "abbrev": canonical,
            "payroll": row[2],
            "soft_cap_space": row[3],
            "luxury_tax_space": row[4],
            "apron1_space": row[5],
            "apron2_space": row[6],
            "g_league_budget": row[10] if len(row) > 10 else None,
        }


def extract_league_office(ws):
    """Pluck cap thresholds for 2026-27 from League Office tab. All in col H (idx 7), rows 2-9."""
    return {
        "salary_floor":   ws.cell(row=2, column=8).value,
        "soft_cap":       ws.cell(row=3, column=8).value,
        "luxury_tax":     ws.cell(row=4, column=8).value,
        "apron1":         ws.cell(row=5, column=8).value,
        "apron2":         ws.cell(row=6, column=8).value,
        "mle":            ws.cell(row=7, column=8).value,
        "tpmle":          ws.cell(row=8, column=8).value,
        "bi_annual":      ws.cell(row=9, column=8).value,
    }


# ---------------------------------------------------------------------------
# SQL generation
# ---------------------------------------------------------------------------

def emit(s=""):
    print(s)


def main():
    if len(sys.argv) < 2:
        sys.stderr.write("usage: import_nba_xlsx.py <path-to-xlsx>\n")
        sys.exit(2)
    path = Path(sys.argv[1])
    if not path.exists():
        sys.stderr.write(f"file not found: {path}\n")
        sys.exit(2)

    sys.stderr.write(f"Loading workbook (this takes a few seconds)...\n")
    wb = openpyxl.load_workbook(path, data_only=True)

    # 1. Extract everything in memory first
    sys.stderr.write("Extracting team rosters...\n")
    rosters = []
    for tab, abbrev in TEAM_TAB_TO_ABBREV.items():
        if tab not in wb.sheetnames:
            sys.stderr.write(f"  WARN: missing tab {tab}\n")
            continue
        for p in extract_team_roster(wb[tab], tab, abbrev):
            rosters.append(p)
    sys.stderr.write(f"  {len(rosters)} player rows from team tabs\n")

    sys.stderr.write("Extracting Player Teams (Fantrax IDs)...\n")
    pt = list(extract_player_teams(wb["Player Teams"]))
    sys.stderr.write(f"  {len(pt)} entries\n")

    sys.stderr.write("Extracting agency assignments...\n")
    aa = list(extract_agency_assignments(wb["Agent Directory"]))
    sys.stderr.write(f"  {len(aa)} player→agency rows\n")

    sys.stderr.write("Extracting team financials...\n")
    fin = list(extract_team_financials(wb["Franchise Financial Status"]))
    sys.stderr.write(f"  {len(fin)} team financial snapshots\n")

    sys.stderr.write("Extracting league office cap data...\n")
    lo = extract_league_office(wb["League Office"])

    # 2. Emit SQL
    emit("-- Generated by scripts/import_nba_xlsx.py")
    emit(f"-- Source: {path.name}")
    emit("BEGIN;")
    emit()

    # 2a. Rename agencies
    emit("-- Rename agencies to spreadsheet-canonical names")
    for sheet_name, uid in AGENCY_NAME_TO_ID.items():
        emit(f"UPDATE agencies SET name = {sql_str(sheet_name)} WHERE id = '{uid}';")
    emit()

    # 2b. League settings (refresh 2026 row)
    emit("-- Refresh league_settings for 2026 from League Office tab")
    emit(f"""UPDATE league_settings SET
    salary_cap        = {sql_num(lo['soft_cap'])},
    luxury_tax_limit  = {sql_num(lo['luxury_tax'])},
    apron_first       = {sql_num(lo['apron1'])},
    apron_second      = {sql_num(lo['apron2'])},
    mle_amount        = {sql_num(lo['mle'])},
    tpmle_amount      = {sql_num(lo['tpmle'])},
    bae_amount        = {sql_num(lo['bi_annual'])}
WHERE league_id = '{LEAGUE_ID}' AND year = 2026;""")
    emit()

    # 2c. Team financials
    emit("-- Refresh team financials from Franchise Financial Status tab")
    for f in fin:
        emit(f"""UPDATE teams SET
    cap_space             = {sql_num(f['soft_cap_space'])},
    luxury_tax_balance    = {sql_num(f['luxury_tax_space'])},
    trade_exception_balance = {sql_num(f['apron1_space'])},
    g_league_budget       = COALESCE({sql_num(f['g_league_budget'])}, g_league_budget)
WHERE league_id = '{LEAGUE_ID}' AND abbreviation = '{f['abbrev']}';""")
    emit()

    # 2d. Players: TRUNCATE + bulk insert
    emit("-- Wipe + re-insert players (xlsx is the source of truth)")
    emit("TRUNCATE players RESTART IDENTITY CASCADE;")
    emit()

    # Build a name → fantrax_id lookup from Player Teams (for backfill)
    name_to_fantrax = {}
    for entry in pt:
        # Use canonical real_life_team if known, else just name
        key_a = (entry["name"].lower(), entry["real_life_team"] or "")
        key_b = entry["name"].lower()
        if entry["fantrax_id"] and entry["real_life_team"]:
            name_to_fantrax[key_a] = entry["fantrax_id"]
        if entry["fantrax_id"] and key_b not in name_to_fantrax:
            name_to_fantrax[key_b] = entry["fantrax_id"]

    # Build a name → agency name lookup
    name_to_agency = {}
    for entry in aa:
        name_to_agency[entry["name"].lower()] = entry["agency"]

    # Insert one player at a time. team_id resolved via subquery on abbreviation.
    # Dedupe by fantrax_id — keep the first occurrence (typically the NBA-roster entry).
    emit("-- Player inserts: team_id resolved via abbreviation subquery, fantrax_id and agency_id from lookup tables.")
    inserted = 0
    skipped_no_team = 0
    skipped_dup = 0
    seen_fantrax = set()
    for p in rosters:
        # Team UUID via subquery
        team_subq = f"(SELECT id FROM teams WHERE league_id = '{LEAGUE_ID}' AND abbreviation = '{p['team_abbrev']}' LIMIT 1)"

        # Fantrax ID lookup
        rl_team = ALL_TEAM_ALIASES.get(p["real_life_team"], p["real_life_team"])
        key_a = (p["name"].lower(), rl_team or "")
        fantrax_id = name_to_fantrax.get(key_a) or name_to_fantrax.get(p["name"].lower())

        # Dedupe: drop subsequent rows for the same Fantrax ID
        if fantrax_id and fantrax_id in seen_fantrax:
            skipped_dup += 1
            continue
        if fantrax_id:
            seen_fantrax.add(fantrax_id)

        # Agency lookup
        agency_sheet_name = name_to_agency.get(p["name"].lower())
        agency_id = AGENCY_NAME_TO_ID.get(agency_sheet_name) if agency_sheet_name else None

        # Build first_name / last_name
        parts = p["name"].split(" ", 1)
        first = parts[0]
        last = parts[1] if len(parts) > 1 else ""

        # Build contract column values
        contract_cols = []
        contract_vals = []
        for year in range(2026, 2041):
            v = p["contracts"].get(year)
            contract_cols.append(f"contract_{year}")
            contract_vals.append(sql_str(v))

        ann_json = "'{}'::jsonb"
        if p["annotations"]:
            import json
            ann_json = sql_str(json.dumps(p["annotations"])) + "::jsonb"

        cols = ["league_id", "team_id", "first_name", "last_name", "position",
                "real_life_team", "fantrax_id", "agency_id",
                "on_two_way", "on_g_league", "on_active_roster",
                "contract_annotations"] + contract_cols
        vals = [
            f"'{LEAGUE_ID}'", team_subq,
            sql_str(first), sql_str(last), sql_str(p["position"]),
            sql_str(p["real_life_team"]), sql_str(fantrax_id),
            f"'{agency_id}'" if agency_id else "NULL",
            "TRUE" if p["on_two_way"] else "FALSE",
            "TRUE" if p["on_g_league"] else "FALSE",
            "TRUE",  # on_active_roster — refine later if needed
            ann_json,
        ] + contract_vals

        emit(f"INSERT INTO players ({', '.join(cols)}) VALUES ({', '.join(vals)});")
        inserted += 1

    emit()
    emit(f"-- Inserted {inserted} players; {skipped_dup} dedup'd by fantrax_id; {skipped_no_team} skipped (no team match)")
    emit()

    # 2e. Backfill agency_id for any unrostered FA in Agent Directory we missed
    # (Already covered above via name_to_agency lookup during insert.)

    emit("COMMIT;")
    emit("-- DONE")

    sys.stderr.write(f"\nDone. {inserted} player INSERTs emitted.\n")


if __name__ == "__main__":
    main()
